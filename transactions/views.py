from django.views.decorators.http import require_POST
from .models import ServiceBillItem

# Receive payment for ServiceBillItem
@require_POST
def receive_payment_service(request):
    from .forms import PaymentForm
    from django.shortcuts import get_object_or_404, redirect
    from django.contrib import messages
    form = PaymentForm(request.POST)
    if form.is_valid():
        servicebill = get_object_or_404(ServiceBillItem, id=form.cleaned_data['servicebill_id'])
        received = form.cleaned_data['amount_received']
        # Update ServiceBillItem amounts
        servicebill.amount_paid += received
        # Update amount_change (balance due)
        servicebill.amount_change = max((servicebill.grand_total or 0) - (servicebill.amount_paid or 0), 0)
        # Update status
        if servicebill.amount_change <= 0:
            servicebill.status = "Paid"
        else:
            servicebill.status = "Balance"
        servicebill.save()
        messages.success(request, f"Payment of ₹{received:.2f} received for service bill #{servicebill.id}.")
    else:
        messages.error(request, "Invalid payment data submitted.")
    return redirect('cashbanklist')
from django.views.decorators.http import require_GET

# API endpoint to get balance for a ServiceBillItem
@require_GET
def service_balance_api(request):
    servicebill_id = request.GET.get('servicebill_id')
    if not servicebill_id:
        return JsonResponse({'error': 'Missing servicebill_id'}, status=400)
    try:
        servicebill = ServiceBillItem.objects.get(id=servicebill_id)
        grand_total = servicebill.grand_total or 0
        amount_paid = servicebill.amount_paid or 0
        balance = grand_total - amount_paid
        return JsonResponse({
            'servicebill_id': servicebill.id,
            'grand_total': float(grand_total),
            'amount_paid': float(amount_paid),
            'balance': float(balance),
            'status': servicebill.status
        })
    except ServiceBillItem.DoesNotExist:
        return JsonResponse({'error': 'ServiceBillItem not found'}, status=404)
# ServiceCreateView for rendering service_create.html
from django.views import View
from django.shortcuts import render

    
# Standard library imports
from functools import reduce
import json
import logging
import operator

# Django core imports
from django.http import JsonResponse, HttpResponse
from django.urls import reverse, reverse_lazy
from django.shortcuts import render
from django.db import transaction
from django.core.paginator import Paginator
from django.views.decorators.csrf import csrf_exempt
import django_tables2 as tables
from django_tables2.export.views import ExportMixin

# Class-based views
from django.views.generic import DetailView, ListView
from django.views.generic.edit import CreateView, UpdateView, DeleteView

# Import Expense views
from .views_expense import ExpenseListView, ExpenseCreateView

# Authentication and permissions
from django.contrib.auth.mixins import LoginRequiredMixin, UserPassesTestMixin

# Third-party packages
from django_tables2 import SingleTableView
from openpyxl import Workbook
from openpyxl.styles import Alignment

# Local app imports
from store.models import Item
from accounts.models import Customer, Vendor
from transactions.tables import PurchasedItemTable
from .models import PurchaseDetail, Sale, Purchase, SaleDetail, Bankaccount, Itempurchased, catogaryitempurchased, ServiceBillItem
from .forms import BankForm
from store.forms import ItemForm
from store.models import catogaryitem 
from django.db.models import Q


logger = logging.getLogger(__name__)


def is_ajax(request):
    return request.META.get('HTTP_X_REQUESTED_WITH') == 'XMLHttpRequest'

def export_bank_to_excel(request):
    # Create a workbook and select the active worksheet.
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = 'Bank Details'

    # Define the column headers
    columns = [
        'ID', 'Account Name', 'Opening Balance', 'As Of Date'
    ]
    worksheet.append(columns)

    # Fetch sales data
    banks = Bankaccount.objects.all()

    for bank in banks:
        # Convert timezone-aware datetime to naive datetime
        if bank.as_of_date.tzinfo is not None:
            as_of_date = bank.as_of_date.replace(tzinfo=None)
        else:
            as_of_date = bank.as_of_date

        worksheet.append([
            bank.id,
            bank.account_name,
            bank.opening_balance,
            as_of_date
        ])

    # Set up the response to send the file
    response = HttpResponse(
        content_type=(
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
    )
    response['Content-Disposition'] = 'attachment; filename=Bank Details.xlsx'
    workbook.save(response)

    return response


from openpyxl import Workbook
from django.http import HttpResponse
from .models import Sale  # Adjust if needed
import datetime

def export_sales_to_excel(request):
    # Create a workbook and sheet
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = 'Sales'

    # Define headers (added 'Item Descriptions')
    columns = [
        'INV no', 'Date', 'Customer', 'Sub Total',
        'Grand Total', 'Tax Amount', 'Tax Percentage',
        'Amount Paid', 'Amount Change', 'Item Descriptions'
    ]
    worksheet.append(columns)

    # Fetch all sales
    sales = Sale.objects.all()

    for sale in sales:
        # Handle datetime conversion
        date_added = sale.date_added.replace(tzinfo=None) if sale.date_added.tzinfo else sale.date_added

        # Generate item descriptions
        items = sale.get_item_descriptions_by_type()
        description_lines = []

        for category in ['PROCESSOR', 'RAM', 'HDD', 'SSD']:
            category_items = items.get(category, [])
            for line in category_items:
                description_lines.append(line.strip())

        description_text = "\n".join(description_lines)  # Line-by-line text

        # Add row data
        row_data = [
            'INV'+ str(sale.id),
            date_added,
            sale.customer.phone,
            sale.sub_total,
            sale.grand_total,
            sale.tax_amount,
            sale.tax_percentage,
            sale.amount_paid,
            sale.amount_change,
            description_text
        ]

        worksheet.append(row_data)

        # Apply wrap text to the "Item Descriptions" column
        description_cell = worksheet.cell(row=worksheet.max_row, column=10)
        description_cell.alignment = Alignment(wrap_text=True)

    # Adjust column widths (optional)
    for col in worksheet.columns:
        max_length = max((len(str(cell.value)) if cell.value else 0) for cell in col)
        worksheet.column_dimensions[col[0].column_letter].width = min(max_length + 2, 50)

    # Prepare response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=sales.xlsx'
    workbook.save(response)
    return response

def export_purchases_to_excel(request):
    # Create a workbook and sheet
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = 'purchases'

    # Define headers (added 'Item Descriptions')
    columns = [
        'INV no', 'Date', 'Vendor', 'Sub Total',
        'Grand Total', 'Tax Amount', 'Tax Percentage',
        'Amount Paid', 'Amount Change', 'Item Descriptions'
    ]
    worksheet.append(columns)

    # Fetch all sales
    purchses = Purchase.objects.all()

    for purchse in purchses:
        # Handle datetime conversion
        date_added = purchse.date_added.replace(tzinfo=None) if purchse.date_added.tzinfo else purchse.date_added

        # Generate item descriptions
        items = purchse.get_item_descriptions_by_type()
        description_lines = []

        for category in ['PROCESSOR', 'RAM', 'HDD', 'SSD']:
            category_items = items.get(category, [])
            for line in category_items:
                description_lines.append(line.strip())

        description_text = "\n".join(description_lines)  # Line-by-line text

        # Add row data
        row_data = [
            'INV'+ str(purchse.id),
            date_added,
            purchse.vendor.name,
            purchse.sub_total,
            purchse.grand_total,
            purchse.tax_amount,
            purchse.tax_percentage,
            purchse.amount_paid,
            purchse.amount_change,
            description_text
        ]

        worksheet.append(row_data)

        # Apply wrap text to the "Item Descriptions" column
        description_cell = worksheet.cell(row=worksheet.max_row, column=10)
        description_cell.alignment = Alignment(wrap_text=True)

    # Adjust column widths (optional)
    for col in worksheet.columns:
        max_length = max((len(str(cell.value)) if cell.value else 0) for cell in col)
        worksheet.column_dimensions[col[0].column_letter].width = min(max_length + 2, 50)

    # Prepare response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=purchases.xlsx'
    workbook.save(response)
    return response


class ServiceCreateView(View):
    def get(self, request, *args, **kwargs):
        # Pass bank accounts for dropdown
        context = {
            "bank_accounts": Bankaccount.objects.all()
        }
        return render(request, 'transactions/service_create.html', context)

    def post(self, request, *args, **kwargs):
        data = request.POST
        customer_id = data.get("customer_id")
        customer = Customer.objects.get(id=customer_id)

        items_json = data.get("items")
        items = json.loads(items_json) if items_json else []

        payment_type = data.get("payment_type", "Cash")
        bank_account_id = data.get("bank_account")
        amount_paid = data.get("amount_paid")
        bank_account_obj = None
        if payment_type == "Bank" and bank_account_id:
            try:
                bank_account_obj = Bankaccount.objects.get(id=bank_account_id)
            except Bankaccount.DoesNotExist:
                bank_account_obj = None

        servicebill_items = []
        for item in items:
            # Calculate amount_change for this item
            try:
                grand_total_val = float(data.get("grand_total", 0))
                amount_paid_val = float(amount_paid or 0)
                amount_change_val = grand_total_val - amount_paid_val
            except Exception:
                amount_change_val = 0
            sb = ServiceBillItem.objects.create(
                customer=customer,
                total_amount=data.get("total_amount"),
                total_tax=data.get("total_tax"),
                grand_total=data.get("grand_total"),
                item_name=item.get("item_name"),
                description=item.get("description"),
                qty=item.get("qty"),
                amount=item.get("amount"),
                tax_percent=item.get("tax_percent"),
                tax_amt=item.get("tax_amt"),
                payment_type=payment_type,
                bank_account=bank_account_obj,
                status="Paid" if amount_paid and float(amount_paid) >= float(data.get("grand_total", 0)) else ("Unpaid" if not amount_paid or float(amount_paid) == 0 else "Balance"),
                amount_paid=amount_paid,
                amount_change=amount_change_val,
            )
            servicebill_items.append(sb)

        # Optionally, update bank balance if payment_type is Bank
        from decimal import Decimal
        if payment_type == "Bank" and bank_account_obj and amount_paid:
            bank_account_obj.opening_balance += Decimal(str(amount_paid))
            bank_account_obj.save()

        if servicebill_items:
            redirect_url = f"/transactions/servicebill/{servicebill_items[0].pk}/invoice/"
            return JsonResponse({"success": True, "redirect_url": redirect_url})
        else:
            return JsonResponse({"success": False, "error": "No items to save."})


from django.core.paginator import Paginator
from django.views.generic import ListView
from .models import Sale, ServiceBillItem

from django.views.generic import ListView
from django.db.models import Q
from .models import Sale, ServiceBillItem

class SaleListView(ListView):
    template_name = 'transactions/sales_list.html'
    context_object_name = 'combined_list'
    paginate_by = 10

    def get_queryset(self):
        # Get all Sales and ServiceBillItems
        sales_qs = Sale.objects.select_related('customer').prefetch_related('saledetail_set__item')
        service_qs = ServiceBillItem.objects.select_related('customer')

        # Get search filters
        status_filter = self.request.GET.get('status', '')
        inv_filter = self.request.GET.get('inv', '').upper()
        serial_filter = self.request.GET.get('serial', '')
        mobileno_filter = self.request.GET.get('mobileno', '')


        # --- Filter Sales ---
        if status_filter:
            sales_qs = sales_qs.filter(status=status_filter)

        if inv_filter:
            if inv_filter.startswith("INV"):
                inv_filter_num = inv_filter[3:]
            else:
                inv_filter_num = inv_filter
            if inv_filter_num.isdigit():
                sales_qs = sales_qs.filter(id=int(inv_filter_num))
            else:
                sales_qs = sales_qs.none()

        if serial_filter:
            sales_qs = sales_qs.filter(saledetail_set__item__serialno__icontains=serial_filter)

        if mobileno_filter:
            sales_qs = sales_qs.filter(customer__phone__icontains=mobileno_filter)

        # --- Filter ServiceBillItems ---
        # Only apply non-status filters to ServiceBillItem
        if inv_filter:
            if inv_filter.startswith("SVC"):
                inv_filter_num = inv_filter[3:]
            else:
                inv_filter_num = inv_filter
            if inv_filter_num.isdigit():
                service_qs = service_qs.filter(id=int(inv_filter_num))
            else:
                service_qs = service_qs.none()

        if serial_filter:
            service_qs = service_qs.filter(description__icontains=serial_filter)  # Or custom field if serial exists

        if mobileno_filter:
            service_qs = service_qs.filter(customer__phone__icontains=mobileno_filter)

        servies1 =  self.request.GET.get('status', '')
        if servies1 == 'Service':
            combined_list = list(service_qs)
        elif servies1 == '':
            combined_list = list(sales_qs) + list(service_qs)
        else:
            combined_list = list(sales_qs)

        # Sort combined list by date_created (Sales might have date_added)
        def get_date(obj):
            if hasattr(obj, 'date_created') and obj.date_created:
                return obj.date_created
            elif hasattr(obj, 'date_added') and obj.date_added:
                return obj.date_added
            return None

        combined_list.sort(key=get_date, reverse=True)

        return combined_list

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['status_filter'] = self.request.GET.get('status', '')
        context['inv_filter'] = self.request.GET.get('inv', '')
        context['serial_filter'] = self.request.GET.get('serial', '')
        context['mobileno_filter'] = self.request.GET.get('mobileno', '')
        # Manual pagination
        page = self.request.GET.get('page', 1)
        paginator = Paginator(self.get_queryset(), self.paginate_by)
        paged_records = paginator.get_page(page)
        context['sales'] = paged_records
        context['is_paginated'] = paged_records.has_other_pages()
        context['page_obj'] = paged_records
        context['paginator'] = paginator
        return context




class SaleDetailView(LoginRequiredMixin, DetailView):
    """
    View to display details of a specific sale.
    """
    model = Sale
    template_name = "transactions/invoice.html"
    context_object_name = "sale"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        sale = self.get_object()
        context["total_after_payment"] = sale.grand_total - sale.amount_paid
        return context


class ServiesDetailView(LoginRequiredMixin, DetailView):
    """
    View to display details of a specific sale.
    """
    model = ServiceBillItem
    template_name = "transactions/servies_inv.html"
    context_object_name = "ServiceBillItem"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        sale = self.get_object()
        context["total_after_payment"] = sale.grand_total - sale.amount_paid
        return context



from django.shortcuts import render
from django.http import JsonResponse
from django.db import transaction
import json
import logging
from .models import Customer, Item, Sale, SaleDetail

logger = logging.getLogger(__name__)


def SaleCreateView(request):
    context = {
        "active_icon": "sales",
        "customers": [c.to_select2() for c in Customer.objects.all()],
        "items": [d.to_select3() for d in Item.objects.all()],
        "bank_accounts": Bankaccount.objects.all()
    }

    if request.method == 'POST' and is_ajax(request=request):
        try:
            data = json.loads(request.body)
            logger.info(f"Received data: {data}")

            # ✅ Add 'payment_type' to required fields
            required_fields = [
                'customer', 'sub_total', 'grand_total',
                'amount_paid', 'amount_change', 'items', 'payment_type'
            ]

            for field in required_fields:
                if field not in data:
                    raise ValueError(f"Missing required field: {field}")

            # ✅ Validate bank account only if payment type is 'Bank'
            bank_account = None
            if data['payment_type'] == 'Bank':
                if 'bank_account' not in data or not data['bank_account']:
                    raise ValueError("Bank account must be provided for Bank payments")
                try:
                    bank_account = Bankaccount.objects.get(id=int(data['bank_account']))
                except Bankaccount.DoesNotExist:
                    raise ValueError("Invalid bank account selected")
                
            # Calculate dynamic status
            amount_paid = float(data["amount_paid"])
            grand_total = float(data["grand_total"])

            if amount_paid >= grand_total:
                sale_status = 'Paid'
            elif amount_paid == 0:
                sale_status = 'Unpaid'
            else:
                sale_status = 'Balance'

            # ✅ Build sale (main) attributes
            sale_attributes = {
                "customer": Customer.objects.get(id=int(data['customer'])),
                "sub_total": float(data["sub_total"]),
                "grand_total": float(data["grand_total"]),
                "tax_amount": float(data.get("tax_amount", 0.0)),
                "tax_percentage": float(data.get("tax_percentage", 0.0)),
                "amount_paid": float(data["amount_paid"]),
                "amount_change": float(data["amount_change"]),
                "payment_type": data['payment_type'],
                "total_discount_amount": data['total_discount_amount'],
                "actual_total_price_before_discount": data['actual_total_price_before_discount'],
                "bank_account": bank_account,
                "status": sale_status  # <-- Now saving dynamic status
            }
            with transaction.atomic():
                # ✅ Create Sale (master)
                new_sale = Sale.objects.create(**sale_attributes)
                logger.info(f"Sale created: {new_sale}")

                # ✅ Process each item
                items = data["items"]
                if not isinstance(items, list):
                    raise ValueError("Items should be a list")

                for item in items:
                    if not all(k in item for k in ["id", "price", "quantity", "total_item"]):
                        raise ValueError("Item is missing required fields")

                    item_instance = Item.objects.get(id=int(item["id"]))

                    if item_instance.quantity < int(item["quantity"]):
                        raise ValueError(f"Not enough stock for item: {item_instance.name}")

                    # ✅ Auto-generate description
                    components = catogaryitem.objects.filter(serial_no=item_instance.serialno)
                    grouped = {
                        'processor': [],
                        'ram': [],
                        'hdd': [],
                        'ssd': []
                    }

                    for comp in components:
                        grouped[comp.category].append(f"{comp.name} X({comp.quantity})")

                    description = ""
                    for cat in ['processor', 'ram', 'hdd', 'ssd']:
                        if grouped[cat]:
                            category_display = cat.upper()
                            values = ", ".join(grouped[cat])
                            description += f"{category_display}: {values}<br>"

                    # ✅ Create SaleDetail with generated description
                    SaleDetail.objects.create(
                        sale=new_sale,
                        item=item_instance,
                        quantity=int(item["quantity"]),
                        price=float(item["price"]),
                        sell_price=float(item["sell_price"]),
                        total_detail=float(item["total_item"]),
                        discount_amount=float(item.get("discount_amount", 0.0)),
                        gst_amount=float(item.get("gst_amount", 0.0)),
                        description=description  # Auto-generated HTML
                    )

                    # ✅ Update stock
                    item_instance.quantity -= int(item["quantity"])
                    item_instance.save()

                
                

            return JsonResponse({
                'status': 'success',
                'message': 'Sale created successfully!',
                'redirect': f'/transactions/sale/{new_sale.pk}/'  # Redirect to invoice page
            })

        except json.JSONDecodeError:
            return JsonResponse({'status': 'error', 'message': 'Invalid JSON format in request body!'}, status=400)
        except Customer.DoesNotExist:
            return JsonResponse({'status': 'error', 'message': 'Customer does not exist!'}, status=400)
        except Item.DoesNotExist:
            return JsonResponse({'status': 'error', 'message': 'Item does not exist!'}, status=400)
        except ValueError as ve:
            return JsonResponse({'status': 'error', 'message': f'Value error: {str(ve)}'}, status=400)
        except TypeError as te:
            return JsonResponse({'status': 'error', 'message': f'Type error: {str(te)}'}, status=400)
        except Exception as e:
            logger.error(f"Exception during sale creation: {e}")
            return JsonResponse({'status': 'error', 'message': f'There was an error during the creation: {str(e)}'}, status=500)

    return render(request, "transactions/sale_create.html", context=context)


# def SaleCreateView(request):
#     context = {
#         "active_icon": "sales",
#         "customers": [c.to_select2() for c in Customer.objects.all()],
#         "items": [d.to_select3() for d in Item.objects.all()]
#     }

#     if request.method == 'POST':
#         if is_ajax(request=request):
#             try:
#                 # Load the JSON data from the request body
#                 data = json.loads(request.body)
#                 print('logger', f"Received data: {data}")
#                 logger.info(f"Received data: {data}")

#                 # Validate required fields
#                 required_fields = [
#                     'customer','item','sub_total', 'grand_total',
#                     'amount_paid', 'amount_change', 'items'
#                 ]
#                 for field in required_fields:
#                     if field not in data:
#                         raise ValueError(f"Missing required field: {field}")

#                 print('tax_amount tax_percentage',float(data.get("tax_amount", 0.0)), float(data.get("tax_percentage", 0.0)))
#                 # Create sale attributes
#                 sale_attributes = {
#                     "customer": Customer.objects.get(id=int(data['customer'])),
#                     "item": Item.objects.get(id=int(data['items'][0]['id'])),
#                     "sub_total": float(data["sub_total"]),
#                     "grand_total": float(data["grand_total"]),
#                     "tax_amount": float(data.get("tax_amount", 0.0)),
#                     "tax_percentage": float(data.get("tax_percentage", 0.0)),
#                     "amount_paid": float(data["amount_paid"]),
#                     "amount_change": float(data["amount_change"]),
#                 }

#                 # Use a transaction to ensure atomicity
#                 with transaction.atomic():
#                     # Create the sale
#                     new_sale = Sale.objects.create(**sale_attributes)
#                     print('Sale created: {new_sale}')
#                     logger.info(f"Sale created: {new_sale}")

#                     # Create sale details and update item quantities
#                     items = data["items"]
#                     if not isinstance(items, list):
#                         raise ValueError("Items should be a list")

#                     for item in items:
#                         if not all(
#                             k in item for k in [
#                                 "id", "price", "quantity", "total_item"
#                             ]
#                         ):
#                             raise ValueError("Item is missing required fields")
                        
#                         print('int(item["id"])', int(item["id"]))
                        

#                         item_instance = Item.objects.get(id=int(item["id"]))
#                         print('item_instance.quantity,  int(item["quantity"])', item_instance.quantity , int(item["quantity"]))
#                         if item_instance.quantity < int(item["quantity"]):
#                             raise ValueError(f"Not enough stock for item: {item_instance.name}")

                        # detail_attributes = {
                        #     "sale": new_sale,
                        #     "item": item_instance,
                        #     "price": float(item["price"]),
                        #     "quantity": int(item["quantity"]),
                        #     "total_detail": float(item["total_item"])
                        # }
#                         SaleDetail.objects.create(**detail_attributes)
#                         logger.info(f"Sale detail created: {detail_attributes}")

#                         # Reduce item quantity
#                         item_instance.quantity -= int(item["quantity"])
#                         item_instance.save()

#                 return JsonResponse(
#                     {
#                         'status': 'success',
#                         'message': 'Sale created successfully!',
#                         'redirect': '/transactions/sales/'
#                     }
#                 )

#             except json.JSONDecodeError:
#                 return JsonResponse(
#                     {
#                         'status': 'error',
#                         'message': 'Invalid JSON format in request body!'
#                     }, status=400)
#             except Customer.DoesNotExist:
#                 return JsonResponse({
#                     'status': 'error',
#                     'message': 'Customer does not exist!'
#                     }, status=400)
#             except Item.DoesNotExist:
#                 return JsonResponse({
#                     'status': 'error',
#                     'message': 'Item does not exist!'
#                     }, status=400)
#             except ValueError as ve:
#                 return JsonResponse({
#                     'status': 'error',
#                     'message': f'Value error: {str(ve)}'
#                     }, status=400)
#             except TypeError as te:
#                 return JsonResponse({
#                     'status': 'error',
#                     'message': f'Type error: {str(te)}'
#                     }, status=400)
#             except Exception as e:
#                 logger.error(f"Exception during sale creation: {e}")
#                 return JsonResponse(
#                     {
#                         'status': 'error',
#                         'message': (
#                             f'There was an error during the creation: {str(e)}'
#                         )
#                     }, status=500)

#     return render(request, "transactions/sale_create.html", context=context)


class SaleDeleteView(LoginRequiredMixin, UserPassesTestMixin, DeleteView):
    """
    View to delete a sale.
    """

    model = Sale
    template_name = "transactions/saledelete.html"

    def get_success_url(self):
        """
        Redirect to the sales list after successful deletion.
        """
        return reverse("saleslist")

    def test_func(self):
        """
        Allow deletion only for superusers.
        """
        return self.request.user.is_superuser


class PurchaseOrderListView(LoginRequiredMixin, ListView):

    """
    View to list all purchases with pagination.
    """

    model = Purchase
    template_name = "transactions/purchase_list.html"
    context_object_name = "purchase"
    paginate_by = 10


    def get_queryset(self):
        sales = Purchase.objects.all().prefetch_related('purchasedetail_set__item')
        queryset = super().get_queryset()
        status = self.request.GET.get('status')
        serial_filter = self.request.GET.get('serial')
        inv = self.request.GET.get('inv')

        if status:
            queryset = queryset.filter(status=status)

        if inv:
            # Assuming invoice number is derived from `id`, e.g., INV123
            if inv.lower().startswith("inv"):
                inv = inv[3:]  # Strip "INV" prefix
            if inv.isdigit():
                queryset = queryset.filter(id=inv)
            else:
                queryset = queryset.none()

        if serial_filter:
            queryset = sales.filter(purchasedetail_set__item__serialno__icontains=serial_filter)

        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['status_filter'] = self.request.GET.get('status', '')
        return context

# class PurchaseCreateListView(LoginRequiredMixin, ListView):
def PurchaseCreateListView(request):
    """
    View to list all purchases with pagination.
    """

    # model = Itempurchased
    # template_name = "transactions/purchased_create.html"
    # context_object_name = "Itempurchased"
    # paginate_by = 10

    context = {
        "active_icon": "purchases",
        "vendors": [c.to_select2() for c in Vendor.objects.all()],
        "items": [d.to_select3() for d in Itempurchased.objects.all()],
        "bank_accounts": Bankaccount.objects.all()
    }

    if request.method == 'POST' and is_ajax(request=request):
        
        try:
            data = json.loads(request.body)
            print('data 1', data)
            logger.info(f"Received data: {data}")

            # ✅ Add 'payment_type' to required fields
            required_fields = [
                'vendor', 'sub_total', 'grand_total',
                'amount_paid', 'amount_change', 'items', 'payment_type'
            ]

            for field in required_fields:
                if field not in data:
                    raise ValueError(f"Missing required field: {field}")

            # ✅ Validate bank account only if payment type is 'Bank'
            bank_account = None
            if data['payment_type'] == 'Bank':
                if 'bank_account' not in data or not data['bank_account']:
                    raise ValueError("Bank account must be provided for Bank payments")
                try:
                    bank_account = Bankaccount.objects.get(id=int(data['bank_account']))
                except Bankaccount.DoesNotExist:
                    raise ValueError("Invalid bank account selected")
                
            # Calculate dynamic status
            amount_paid = float(data["amount_paid"])
            grand_total = float(data["grand_total"])

            if amount_paid >= grand_total:
                purchase_status = 'Paid'
            elif amount_paid == 0:
                purchase_status = 'Unpaid'
            else:
                purchase_status = 'Balance'

            # ✅ Build sale (main) attributes
            sale_attributes = {
                "vendor": Vendor.objects.get(id=int(data['vendor'])),
                "sub_total": float(data["sub_total"]),
                "grand_total": float(data["grand_total"]),
                "tax_amount": float(data.get("tax_amount", 0.0)),
                "tax_percentage": float(data.get("tax_percentage", 0.0)),
                "amount_paid": float(data["amount_paid"]),
                "amount_change": float(data["amount_change"]),
                "payment_type": data['payment_type'],
                "bank_account": bank_account,
                "status": purchase_status  # <-- Now saving dynamic status
            }
            with transaction.atomic():
                # ✅ Create Sale (master)
                new_sale = Purchase.objects.create(**sale_attributes)
                logger.info(f"Purchase created: {new_sale}")

                # ✅ Process each item
                items = data["items"]
                if not isinstance(items, list):
                    raise ValueError("Items should be a list")

                for item in items:
                    if not all(k in item for k in ["id", "price", "quantity", "total_item"]):
                        raise ValueError("Item is missing required fields")

                    print('int(item["id"]', int(item["id"]))
                    item_instance = Itempurchased.objects.get(id=int(item["id"]))

                    if item_instance.quantity < int(item["quantity"]):
                        raise ValueError(f"Not enough stock for item: {item_instance.name}")

                    # ✅ Auto-generate description
                    components = catogaryitempurchased.objects.filter(serial_no=item_instance.serialno)
                    grouped = {
                        'processor': [],
                        'ram': [],
                        'hdd': [],
                        'ssd': []
                    }

                    for comp in components:
                        grouped[comp.category].append(f"{comp.name} X({comp.quantity})")

                    description = ""
                    for cat in ['processor', 'ram', 'hdd', 'ssd']:
                        if grouped[cat]:
                            category_display = cat.upper()
                            values = ", ".join(grouped[cat])
                            description += f"{category_display}: {values} ({comp.serial_no})<br>"

                    # ✅ Create SaleDetail with generated description
                    PurchaseDetail.objects.create(
                        purchase=new_sale,
                        item=item_instance,
                        quantity=int(item["quantity"]),
                        price=float(item["price"]),
                        total_detail=float(item["total_item"]),
                        description=description  # Auto-generated HTML
                    )

                    # ✅ Update stock
                    item_instance.quantity -= int(item["quantity"])
                    item_instance.save()

                
                

            return JsonResponse({
                'status': 'success',
                'message': 'Purchase created successfully!',
                'redirect': '/transactions/purchases/'
            })

        except json.JSONDecodeError:
            return JsonResponse({'status': 'error', 'message': 'Invalid JSON format in request body!'}, status=400)
        except Customer.DoesNotExist:
            return JsonResponse({'status': 'error', 'message': 'Customer does not exist!'}, status=400)
        except Item.DoesNotExist:
            return JsonResponse({'status': 'error', 'message': 'Item does not exist!'}, status=400)
        except ValueError as ve:
            return JsonResponse({'status': 'error', 'message': f'Value error: {str(ve)}'}, status=400)
        except TypeError as te:
            return JsonResponse({'status': 'error', 'message': f'Type error: {str(te)}'}, status=400)
        except Exception as e:
            logger.error(f"Exception during sale creation: {e}")
            return JsonResponse({'status': 'error', 'message': f'There was an error during the creation: {str(e)}'}, status=500)

    return render(request, "transactions/purchased_create.html", context=context)


class PurchaseListView(LoginRequiredMixin, ExportMixin, tables.SingleTableView):
    """
    View class to update product information.

    Attributes:
    - model: The model associated with the view.
    - template_name: The HTML template used for rendering the view.
    - fields: The fields to be updated.
    - success_url: The URL to redirect to upon successful form submission.
    """

    model = Item
    table_class = PurchasedItemTable
    context_object_name = "Itempurchased"
    paginate_by = 10
    SingleTableView.table_pagination = False

    def get_template_names(self):

        user_profile = self.request.user.profile
        if user_profile.role == 'OP':
            return ["store/operative_dashboard.html"]
        else:
            return ["transactions/purchaselist.html"]
        
    def get_queryset(self):
        # Get base queryset
        queryset = super().get_queryset()
        # Filter items where purchased_type == 'vendor'
        queryset = queryset.filter(purchased_type='vendor')
        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        items = context['Itempurchased']
        # Prefetch vendor data to avoid N+1 queries
        context['vendors'] = Customer.objects.all().order_by('first_name')

        # Get all category items in one go to reduce DB hits
        all_category_items = catogaryitem.objects.filter(
            serial_no__in=[item.serialno for item in items]
        )

        # Group them by serial number, but only store the component name
        from collections import defaultdict
        grouped_data = defaultdict(lambda: {'processors': [], 'rams': [], 'hdds': [], 'ssds': []})

        for data in all_category_items:
            serial = data.serial_no
            if data.category == 'processor':
                grouped_data[serial]['processors'].append(data.name + 'X' + '(' + str(data.quantity) + ')' )
            elif data.category == 'ram':
                grouped_data[serial]['rams'].append(data.name + 'X' + '(' + str(data.quantity) + ')' )
            elif data.category == 'hdd':
                grouped_data[serial]['hdds'].append(data.name +'X' + '(' + str(data.quantity) + ')' )
            elif data.category == 'ssd':
                grouped_data[serial]['ssds'].append(data.name + 'X' + '(' + str(data.quantity) + ')' )

        # Attach to each item
        for item in items:
            serial = item.serialno
            item.processors = grouped_data[serial]['processors']
            item.rams = grouped_data[serial]['rams']
            item.hdds = grouped_data[serial]['hdds']
            item.ssds = grouped_data[serial]['ssds']

        return context





class PurchaseDetailView(LoginRequiredMixin, DetailView):
    """
    View to display details of a specific purchase.
    """

    model = Purchase
    template_name = "transactions/invoice_purchase.html"
    context_object_name = "purchase"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        sale = self.get_object()
        context["total_after_payment"] = sale.grand_total - sale.amount_paid
        return context


from operator import and_

class PurchaseItemSearchListView(PurchaseListView):
    """
    View class to search and display a filtered list of items.

    Attributes:
    - paginate_by: Number of items per page for pagination.
    """
    table_class = PurchasedItemTable
    paginate_by = 10
    SingleTableView.table_pagination = False

    def get_queryset(self):
        result = super().get_queryset()

        query = self.request.GET.get("q")
        vendor = self.request.GET.get("vendor", "").strip()
        purchased_code = self.request.GET.get("purchased_code", "").strip()

        print('jihkjdkhd', query, vendor, purchased_code)

        # if not query or vendor or purchased_code:
        #     return result.none()
        
        if query:
            print('uyuyu')
            query_list = query.split()
            result = result.filter(
                reduce(
                    operator.and_, (Q(name__icontains=q) | Q(serialno__icontains=q) | Q(make_and_models__icontains=q) for q in query_list)
                )
            )

        if vendor:
            query_list = vendor.split()
            result = result.filter(
                reduce(
                    and_,
                        (Q(customer__first_name__icontains=q) | Q(customer__last_name__icontains=q) for q in query_list)
                )
            )
            

        if purchased_code:
            print('iopioii')
            query_list = purchased_code.split()
            result = result.filter(
                reduce(
                    operator.and_, (Q(purchased_code__icontains=q) for q in query_list)
                )
            )
        return result
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        items = context['Itempurchased']  # assuming your ListView or similar sets this

        # Get all category items in one go to reduce DB hits     
        all_category_items = catogaryitem.objects.filter(
            serial_no__in=[item.serialno for item in items]
        )

        # Group them by serial number, but only store the component name
        from collections import defaultdict
        grouped_data = defaultdict(lambda: {'processors': [], 'rams': [], 'hdds': [], 'ssds': []})

        for data in all_category_items:
            serial = data.serial_no

            # Store only the component name, not the category
            if data.category == 'processor':
                grouped_data[serial]['processors'].append(data.name + 'X' + '(' + str(data.quantity) + ')' )  # Store only the name
            elif data.category == 'ram':  # Store only the name
                grouped_data[serial]['rams'].append(data.name + 'X' + '(' + str(data.quantity) + ')' )  # Store only the name
            elif data.category == 'hdd':
                grouped_data[serial]['hdds'].append(data.name +'X' + '(' + str(data.quantity) + ')' )  # Store only the name
            elif data.category == 'ssd':
                grouped_data[serial]['ssds'].append(data.name + 'X' + '(' + str(data.quantity) + ')' )  # Store only the name

        # Attach to each item
        for item in items:
            serial = item.serialno
            item.processors = grouped_data[serial]['processors'] 
            item.rams = grouped_data[serial]['rams']
            item.hdds = grouped_data[serial]['hdds']
            item.ssds = grouped_data[serial]['ssds']

        return context


# class PurchaseCreateView(LoginRequiredMixin, CreateView):
#     """
#     View to create a new purchase.
#     """

#     model = Purchase
#     form_class = PurchaseForm
#     template_name = "transactions/purchases_form.html"

#     def get_success_url(self):
#         """
#         Redirect to the purchases list after successful form submission.
#         """
#         return reverse("purchaseslist")
    


class PurchasedCreateView(LoginRequiredMixin, CreateView):
    model = Item
    template_name = "transactions/purchasecreate.html"
    form_class = ItemForm

    def form_valid(self, form):

        item = form.save(commit=False)
        item.created_by = self.request.user  # Set the user who created the item
        item.purchased_type = 'vendor'
        item.save()
        serialno = item.serialno.strip()


        component_fields = ['processor', 'ram', 'hdd', 'ssd']

        try:
            for component in component_fields:
                names_str = self.request.POST.get(component, '')
                qtys_str = self.request.POST.get(f"{component}_qty", '')

                names = [name.strip() for name in names_str.split(',') if name.strip()]
                qtys = [qty.strip() for qty in qtys_str.split(',') if qty.strip()]

                if len(names) != len(qtys):
                    messages.error(self.request, f"Mismatch in count of names and quantities for {component}.")
                    continue

                for name, qty in zip(names, qtys):
                    try:
                        qty_int = int(qty)
                        if qty_int <= 0:
                            continue
                    except ValueError:
                        messages.error(self.request, f"Invalid quantity '{qty}' for {component} '{name}'")
                        continue

                    catogaryitem.objects.create(
                        name=name.upper(),
                        category=component,
                        serial_no=serialno,
                        quantity=qty_int,
                        unit_price=0.00,  # Set as needed
                        purchased_code = item.purchased_code.strip(),
                        created_by=self.request.user
                    )
        except Exception as e:
            messages.error(self.request, f"Error while saving components: {str(e)}")

        messages.success(self.request, "Purchased and components saved.")
        return super().form_valid(form)

    def get_success_url(self):
        return reverse_lazy('purchaseslist')


# class PurchaseUpdateView(LoginRequiredMixin, UpdateView):
#     """
#     View to update an existing purchase.
#     """
#     model = Purchase
#     form_class = PurchaseForm
#     template_name = "transactions/purchases_form.html"

#     def get_success_url(self):
#         """
#         Redirect to the purchases list after successful form submission.
#         """
#         return reverse("purchaseslist")


class PurchaseDeleteView(LoginRequiredMixin, UserPassesTestMixin, DeleteView):
    """
    View to delete a purchase.
    """

    model = Purchase
    template_name = "transactions/purchasedelete.html"

    def get_success_url(self):
        """
        Redirect to the purchases list after successful deletion.
        """
        return reverse("purchaseslist")

    def test_func(self):
        """
        Allow deletion only for superusers.
        """
        return self.request.user.is_superuser
    
class BankDeleteView(LoginRequiredMixin, UserPassesTestMixin, DeleteView):
    """
    View to delete a Bank.
    """

    model = Bankaccount
    template_name = "transactions/bankdelete.html"

    def get_success_url(self):
        """
        Redirect to the Bank list after successful deletion.
        """
        return reverse("cashbanklist")

    def test_func(self):
        """
        Allow deletion only for superusers.
        """
        return self.request.user.is_superuser
    
from itertools import chain
from operator import attrgetter

class cashbankListView(LoginRequiredMixin, ListView):

    template_name = "transactions/bank_acc.html"
    paginate_by = None

    def get_queryset(self):
        # Not used, but required by ListView
        return Bankaccount.objects.none()

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        # Provide all bank accounts for sidebar
        context['all_bankaccounts'] = Bankaccount.objects.all().order_by('id')

        account_id = self.request.GET.get('account_id')

        # Fetch related transactions
        sales = Sale.objects.select_related('customer').all()
        purchases = Purchase.objects.select_related('vendor').all()
        services = ServiceBillItem.objects.select_related('customer').all()

        if account_id:
            sales = sales.filter(bank_account_id=account_id)
            purchases = purchases.filter(bank_account_id=account_id)
            services = services.filter(bank_account_id=account_id)

        # Add transaction type tag safely
        for s in sales:
            setattr(s, 'transaction_type', "Sale")
        for p in purchases:
            setattr(p, 'transaction_type', "Purchase")
        for sv in services:
            setattr(sv, 'transaction_type', "Service")

        # Only keep objects with a valid date for sorting
        def get_sort_date(obj):
            date = getattr(obj, 'date_added', None) or getattr(obj, 'date_created', None)
            return date if date is not None else ''

        merged_transactions = [t for t in chain(sales, purchases, services) if get_sort_date(t)]
        merged_transactions.sort(key=get_sort_date, reverse=True)

        


        
        # No pagination: show all records
        context['selected_account_id'] = account_id
        context['page_obj'] = merged_transactions
        return context





class BankCreateView(LoginRequiredMixin, CreateView):
    model = Bankaccount
    template_name = "transactions/createbankacc.html"
    form_class = BankForm
        
    def get_success_url(self):
        """
        Redirect to the Bank list after successful deletion.
        """
        return reverse("cashbanklist")

    def test_func(self):
        """
        Allow deletion only for superusers.
        """
        return self.request.user.is_superuser
    

from django.shortcuts import redirect, get_object_or_404
from django.contrib import messages
from django.views.decorators.http import require_POST
from .models import Sale
from .forms import PaymentForm

@require_POST
def receive_payment(request):
    form = PaymentForm(request.POST)
    if form.is_valid():
        sale = get_object_or_404(Sale, id=form.cleaned_data['sale_id'])
        received = form.cleaned_data['amount_received']
        
        # Update Sale amounts
        sale.amount_paid += received
        # Calculate remaining balance
        from decimal import Decimal
        remaining = sale.grand_total - sale.amount_paid

        # Update amount_change (balance due)
        sale.amount_change = max(remaining, Decimal('0.00'))

        # Update status
        if remaining <= 0:
            sale.status = "Paid"
        else:
            sale.status = "Balance"  # Or "Unpaid" depending on your logic

        sale.save()

        messages.success(request, f"Payment of ₹{received:.2f} received for sale #{sale.id}.")
    else:
        messages.error(request, "Invalid payment data submitted.")
    
    return redirect('cashbanklist')  # Replace with your actual list view URL name


@require_POST
def receive_payment_purchase(request):
    form = PaymentForm(request.POST)
    if form.is_valid():
        purchase = get_object_or_404(Purchase, id=form.cleaned_data['purchase_id'])
        received = form.cleaned_data['amount_received']
        
        # Update Sale amounts
        purchase.amount_paid += received
        # Calculate remaining balance
        remaining = purchase.grand_total - purchase.amount_paid
        
        # Update amount_change (balance due)
        purchase.amount_change = max(remaining, 0)
        
        # Update status
        if remaining <= 0:
            purchase.status = "Paid"
        else:
            purchase.status = "Balance"  # Or "Unpaid" depending on your logic
        
        purchase.save()
        
        messages.success(request, f"Payment of ₹{received:.2f} received for purchase #{purchase.id}.")
    else:
        messages.error(request, "Invalid payment data submitted.")
    
    return redirect('cashbanklist')  # Replace with your actual list view URL name



@csrf_exempt
def search_suggestions_purchase(request):
    query = request.GET.get('q', '').strip()
    vendor_query = request.GET.get('vendor', '').strip()
    code_query = request.GET.get('purchased_code', '').strip()

    suggestions = []

    # Start building the Q object dynamically
    filters = Q()

    if query:
        filters |= Q(name__icontains=query)
        filters |= Q(serialno__icontains=query)
        filters |= Q(make_and_models__icontains=query)

    if vendor_query:
        filters |= Q(vendor__name__icontains=vendor_query)

    if code_query:
        filters |= Q(purchase_code__icontains=code_query)

    if filters:
        items = Itempurchased.objects.filter(filters)
        for item in items:
            if query and query.lower() in item.name.lower():
                suggestions.append(item.name)
            elif query and query.lower() in item.serialno.lower():
                suggestions.append(item.serialno)
            elif query and query.lower() in item.make_and_models.lower():
                suggestions.append(item.make_and_models)
            elif vendor_query and item.vendor and vendor_query.lower() in item.vendor.name.lower():
                suggestions.append(item.vendor.name)
            elif code_query and code_query.lower() in item.purchase_code.lower():
                suggestions.append(item.purchase_code)

    suggestions = list(set(suggestions))  # Remove duplicates
    return JsonResponse({'suggestions': suggestions})



def customer_search(request):
    query = request.GET.get('q', '')
    results = []

    if query:
        customers = Customer.objects.filter(
            Q(first_name__icontains=query) | Q(last_name__icontains=query) | Q(phone__icontains=query)
        )
        results = [
            {
                'id': c.id,
                'text': f"{c.first_name} {c.last_name}".strip(),
                'phone': f"{c.phone}".strip()
            } for c in customers
        ]
    return JsonResponse({'results': results})



@require_POST
def customer_create(request):
    first_name = request.POST.get('first_name')
    last_name = request.POST.get('last_name')
    person_type = request.POST.get('person_type', 'customer')  # Default to customer if not specified
    phone = request.POST.get('phone')
    
    if not first_name:
        return JsonResponse({'status': 'error', 'message': 'First name is required.'}, status=400)

    customer = Customer.objects.create(
        first_name=first_name,
        last_name=last_name,
        phone=phone,
        person_type=person_type
    )
    return JsonResponse({'success': True, 'customer': {'id': customer.id, 'name': customer.get_full_name()}})

    # return JsonResponse({'id': customer.id, 'text': customer.get_full_name()})    


def vendor_search(request):
    query = request.GET.get('q', '')
    results = []

    if query:
        vendors = Vendor.objects.filter(
            Q(name__icontains=query) | Q(phone_number__icontains=query)
        )
        results = [
            {
                'id': c.id,
                'text': f"{c.name}".strip(),
                'phone': f"{c.phone_number}".strip()
            } for c in vendors
        ]
    return JsonResponse({'results': results})



@require_POST
def vendor_create(request):
    name = request.POST.get('name')
    phone_number = request.POST.get('phone_number')
    # if not last_name:
    #     return JsonResponse({'status': 'error', 'message': 'First name is required.'}, status=400)

    vendor = Vendor.objects.create(name=name, phone_number = phone_number)
    return JsonResponse({'id': vendor.id, 'text': vendor.get_full_name()})


from django.http import JsonResponse
from django.db import transaction
from accounts.models import Customer
from .models import ServiceBillItem

import json
from django.views import View
from django.shortcuts import render, redirect
from .models import ServiceBillItem, Customer

from django.http import JsonResponse

class ServiceBillCreateView(View):
    def post(self, request, *args, **kwargs):
        data = request.POST
        customer_id = data.get("customer_id")
        customer = Customer.objects.get(id=customer_id)

        # ✅ Add 'payment_type' to required fields
        # required_fields = [
        #     'customer', 'sub_total', 'grand_total',
        #     'amount_paid', 'amount_change', 'items', 'payment_type'
        # ]

        # for field in required_fields:
        #     if field not in data:
        #         raise ValueError(f"Missing required field: {field}")

        # ✅ Validate bank account only if payment type is 'Bank'
        bank_account = None
        if data['payment_type'] == 'Bank':
            if 'bank_account' not in data or not data['bank_account']:
                raise ValueError("Bank account must be provided for Bank payments")
            try:
                bank_account = Bankaccount.objects.get(id=int(data['bank_account']))
            except Bankaccount.DoesNotExist:
                raise ValueError("Invalid bank account selected")
            
        # Calculate dynamic status
        amount_paid = float(data["amount_paid"])
        grand_total = float(data["grand_total"])

        if amount_paid >= grand_total:
            sale_status = 'Paid'
        elif amount_paid == 0:
            sale_status = 'Unpaid'
        else:
            sale_status = 'Balance'

        items_json = data.get("items")
        items = json.loads(items_json) if items_json else []

        servicebill_items = []
        total_received = 0
        for item in items:
            try:
                item_amount = float(item.get("amount") or 0)
            except Exception:
                item_amount = 0
            total_received += item_amount
        try:
            grand_total_val = float(data.get("grand_total", 0))
        except Exception:
            grand_total_val = 0
        amount_change_val = grand_total_val - float(data.get("amount_paid", 0))

        # Create the main bill summary item (first item, or a summary row)
        if items:
            first_item = items[0]
            sb = ServiceBillItem.objects.create(
                customer=customer,
                total_amount=data.get("total_amount"),
                total_tax=data.get("total_tax"),
                grand_total=data.get("grand_total"),
                item_name=first_item.get("item_name"),
                description=first_item.get("description"),
                qty=first_item.get("qty"),
                amount=first_item.get("amount"),
                tax_percent=first_item.get("tax_percent"),
                tax_amt=first_item.get("tax_amt"),
                payment_type=data.get("payment_type"),
                bank_account=bank_account,
                status=sale_status,
                amount_paid=float(data.get("amount_paid", 0)),
                amount_change=amount_change_val
            )
            servicebill_items.append(sb)
            # Add the rest of the items as detail rows (amount_paid and amount_change = 0)
            for item in items[1:]:
                ServiceBillItem.objects.create(
                    customer=customer,
                    total_amount=None,
                    total_tax=None,
                    grand_total=None,
                    item_name=item.get("item_name"),
                    description=item.get("description"),
                    qty=item.get("qty"),
                    amount=item.get("amount"),
                    tax_percent=item.get("tax_percent"),
                    tax_amt=item.get("tax_amt"),
                    payment_type=data.get("payment_type"),
                    bank_account=bank_account,
                    status=sale_status,
                    amount_paid=0,
                    amount_change=0
                )
            redirect_url = f"/transactions/servicebill/{sb.pk}/invoice/"
            return JsonResponse({"success": True, "redirect_url": redirect_url})
        else:
            return JsonResponse({"success": False, "error": "No items to save."})



class ServiceBillInvoiceView(View):
    def get(self, request, pk, *args, **kwargs):
        # Get the first item
        servicebill = get_object_or_404(ServiceBillItem, pk=pk)

        # Fetch all items from same bill (same customer & grand_total & date)
        items = ServiceBillItem.objects.filter(
            customer=servicebill.customer,
            date_created=servicebill.date_created,
            grand_total=servicebill.grand_total
        )


        # Calculate balance
        try:
            grand_total = servicebill.grand_total or 0
            amount_paid = servicebill.amount_paid or 0
            balance = grand_total - amount_paid
        except Exception:
            balance = 0

        context = {
            "servicebill": servicebill,
            "items": items,
            "balance": balance,
        }
        return render(request, "transactions/servies_inv.html", context)


